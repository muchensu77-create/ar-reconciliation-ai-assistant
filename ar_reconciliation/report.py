from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .ai_notes import build_control_checklist, build_exception_memo, build_safe_ai_prompt
from .core import ReconciliationResult


def write_excel_report(result: ReconciliationResult, output_path: str | Path) -> Path:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    summary_df = pd.DataFrame([result.summary])
    memo = build_exception_memo(
        result.summary,
        result.unmatched_invoices,
        result.unmatched_receipts,
        result.exception_queue,
    )
    prompt = build_safe_ai_prompt(result.summary, result.aging)
    ai_notes = pd.DataFrame({"section": ["exception_memo", "safe_ai_prompt"], "content": [memo, prompt]})
    controls = build_control_checklist(result.summary)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="summary", index=False)
        result.matches.to_excel(writer, sheet_name="matches", index=False)
        result.exception_queue.to_excel(writer, sheet_name="exception_queue", index=False)
        result.unmatched_invoices.to_excel(writer, sheet_name="unmatched_invoices", index=False)
        result.unmatched_receipts.to_excel(writer, sheet_name="unmatched_receipts", index=False)
        result.aging.to_excel(writer, sheet_name="aging", index=False)
        result.data_quality.to_excel(writer, sheet_name="data_quality", index=False)
        controls.to_excel(writer, sheet_name="control_checklist", index=False)
        ai_notes.to_excel(writer, sheet_name="ai_notes", index=False)

        for sheet in writer.book.worksheets:
            _format_sheet(sheet)

    return output


def _format_sheet(sheet) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    high_fill = PatternFill(fill_type="solid", fgColor="F8D7DA")
    medium_fill = PatternFill(fill_type="solid", fgColor="FFF3CD")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        width = min(max(max_length + 2, 12), 48)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width

    headers = [cell.value for cell in sheet[1]]
    if "priority" in headers:
        priority_col = headers.index("priority") + 1
        for row in range(2, sheet.max_row + 1):
            value = sheet.cell(row=row, column=priority_col).value
            fill = high_fill if value == "high" else medium_fill if value == "medium" else None
            if fill:
                for col in range(1, sheet.max_column + 1):
                    sheet.cell(row=row, column=col).fill = fill

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, (int, float)) and ("amount" in str(sheet.cell(row=1, column=cell.column).value).lower()):
                cell.number_format = '#,##0.00'

    sheet.freeze_panes = "A2"
